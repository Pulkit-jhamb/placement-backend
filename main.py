from flask import Flask, request, jsonify, make_response, send_file
from flask_cors import CORS
from flask_socketio import SocketIO, emit, join_room, leave_room
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
from dotenv import load_dotenv
import jwt
import os
import uuid
from google import genai
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
import PyPDF2
import docx
import re
import requests

# Import database collections
from database import (
    users_collection,
    messages_collection,
    conversations_collection,
    notes_collection,
    favorites_collection,
    help_reports_collection,
    quizzes_collection,
    quiz_answers_collection,
    quiz_results_collection,
    placement_collection,
    project_collection,  # Student's personal projects
    research_collection,  # Student's personal research
    admin_projects_collection,  # Admin opportunities
    admin_research_collection,  # Admin opportunities
    admin_patents_collection,  # Admin opportunities
    student_applications_collection,  # NEW: Student applications
    otp_collection
)

# Load environment variables
load_dotenv()
SECRET_KEY = os.getenv("SECRET_KEY", "secret123")

# Configure Gemini API
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
if GEMINI_API_KEY:
    print("✅ Gemini API key found")
    # Set the API key as environment variable for genai.Client()
    os.environ["GEMINI_API_KEY"] = GEMINI_API_KEY
else:
    print("⚠️ WARNING: GEMINI_API_KEY not found in environment variables")

# Initialize Flask app
app = Flask(__name__)
app.config["SECRET_KEY"] = SECRET_KEY

# Enable CORS
CORS(
    app,
    resources={
        r"/api/*": {
            "origins": ["http://localhost:5173", "http://localhost:5174"],
            "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
            "allow_headers": ["Content-Type", "Authorization"],
            "expose_headers": ["Content-Range", "X-Content-Range"],
            "supports_credentials": True,
            "max_age": 600
        }
    }
)

# Initialize Socket.IO
socketio = SocketIO(
    app,
    cors_allowed_origins=["http://localhost:5173", "http://localhost:5174"],
    async_mode='threading'
)

# =====================================================
# 🔧 HELPER FUNCTIONS
# =====================================================

def generate_token(user):
    payload = {
        "user_id": str(user["_id"]),
        "email": user["email"],
        "userType": user["userType"],
        "exp": datetime.utcnow() + timedelta(hours=2)
    }
    return jwt.encode(payload, app.config["SECRET_KEY"], algorithm="HS256")

def get_current_user(token=None):
    """Get current user from JWT token"""
    if not token and "Authorization" in request.headers:
        auth_header = request.headers["Authorization"]
        try:
            token = auth_header.split(" ")[1]
        except IndexError:
            return None
            
    if not token:
        return None
        
    try:
        payload = jwt.decode(token, app.config["SECRET_KEY"], algorithms=["HS256"])
        user = users_collection.find_one({"_id": payload["user_id"]})
        return user
    except (jwt.ExpiredSignatureError, jwt.InvalidTokenError) as e:
        print(f"❌ Token error: {e}")
        return None
    except Exception as e:
        print(f"❌ Error getting current user: {e}")
        return None

def validate_drive_link(link):
    """Validate Google Drive link"""
    if not link:
        return False
    return link.startswith("https://drive.google.com/") or link.startswith("https://docs.google.com/")

def extract_text_from_pdf(file):
    """Extract text from PDF file"""
    pdf_reader = PyPDF2.PdfReader(file)
    text = ""
    for page in pdf_reader.pages:
        text += page.extract_text() + "\n"
    return text

def extract_text_from_docx(file):
    """Extract text from DOCX file"""
    doc = docx.Document(file)
    text = ""
    for paragraph in doc.paragraphs:
        text += paragraph.text + "\n"
    return text

def parse_resume(text):
    """Parse resume text and extract structured information"""
    lines = text.split("\n")

    section_keywords = {
        "personal_info": ["name", "contact", "email", "phone", "address", "profile"],
        "education": ["education", "academic", "qualification", "degree", "university", "college"],
        "experience": ["experience", "employment", "work history", "professional experience"],
        "projects": ["projects", "personal projects", "academic projects"],
        "skills": ["skills", "technical skills", "competencies", "expertise"],
        "certifications": ["certifications", "certificates", "licenses"],
        "research": ["research", "research papers", "publications"],
        "thesis": ["thesis", "dissertation"],
        "startups": ["startup", "entrepreneurship", "venture"],
        "achievements": ["achievements", "awards", "honors", "accomplishments"],
        "languages": ["languages", "language proficiency"],
        "interests": ["interests", "hobbies"],
    }

    parsed_data = {
        "personal_info": [],
        "education": [],
        "experience": [],
        "projects": [],
        "skills": [],
        "certifications": [],
        "research": [],
        "thesis": [],
        "startups": [],
        "achievements": [],
        "languages": [],
        "interests": [],
        "other_sections": [],
    }

    current_section = None
    current_content = []

    email_pattern = r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b"
    phone_pattern = r"[\+]?[(]?[0-9]{1,4}[)]?[-\s\.]?[(]?[0-9]{1,4}[)]?[-\s\.]?[0-9]{1,4}[-\s\.]?[0-9]{1,9}"

    all_text = " ".join(lines)
    emails = re.findall(email_pattern, all_text)
    phones = re.findall(phone_pattern, all_text)

    if emails:
        parsed_data["personal_info"].append(f"Email: {emails[0]}")
    if phones:
        parsed_data["personal_info"].append(f"Phone: {phones[0]}")

    for line in lines:
        line = line.strip()
        if not line:
            continue

        is_header = False
        line_lower = line.lower()

        for section_key, keywords in section_keywords.items():
            for keyword in keywords:
                if keyword in line_lower and len(line) < 50:
                    if current_section and current_content:
                        parsed_data[current_section].extend(current_content)

                    current_section = section_key
                    current_content = []
                    is_header = True
                    break
            if is_header:
                break

        if not is_header and current_section:
            current_content.append(line)
        elif not is_header and not current_section:
            parsed_data["personal_info"].append(line)

    if current_section and current_content:
        parsed_data[current_section].extend(current_content)

    parsed_data = {k: v for k, v in parsed_data.items() if v}

    return parsed_data

def calculate_ats_score(parsed_data, full_text):
    """Calculate ATS (Applicant Tracking System) score for the resume"""
    score = 0
    max_score = 100
    feedback = []
    scoring_breakdown = {}

    contact_score = 0
    has_email = any("email" in item.lower() for item in parsed_data.get("personal_info", []))
    has_phone = any("phone" in item.lower() for item in parsed_data.get("personal_info", []))

    if has_email:
        contact_score += 5
        feedback.append("✓ Email address found")
    else:
        feedback.append("✗ Missing email address")

    if has_phone:
        contact_score += 5
        feedback.append("✓ Phone number found")
    else:
        feedback.append("✗ Missing phone number")

    scoring_breakdown["Contact Information"] = {"score": contact_score, "max": 10}
    score += contact_score

    education_score = 0
    education_items = parsed_data.get("education", [])
    if education_items:
        education_score = min(15, len(education_items) * 5)
        feedback.append(f"✓ Education section found with {len(education_items)} entries")
    else:
        feedback.append("✗ No education section found")

    scoring_breakdown["Education"] = {"score": education_score, "max": 15}
    score += education_score

    experience_score = 0
    experience_items = parsed_data.get("experience", [])
    if experience_items:
        experience_score = min(25, len(experience_items) * 5)
        feedback.append(f"✓ Work experience section found with {len(experience_items)} entries")

        has_metrics = any(bool(re.search(r"\d+[%+]?", item)) for item in experience_items)
        if has_metrics:
            experience_score += 5
            feedback.append("✓ Quantifiable achievements found (numbers/metrics)")
    else:
        feedback.append("✗ No work experience section found")

    scoring_breakdown["Work Experience"] = {"score": min(experience_score, 25), "max": 25}
    score += min(experience_score, 25)

    skills_score = 0
    skills_items = parsed_data.get("skills", [])
    if skills_items:
        skills_count = sum(len(item.split(",")) for item in skills_items)
        skills_score = min(15, skills_count)
        feedback.append(f"✓ Skills section found with approximately {skills_count} skills")
    else:
        feedback.append("✗ No skills section found")

    scoring_breakdown["Skills"] = {"score": skills_score, "max": 15}
    score += skills_score

    projects_score = 0
    projects_items = parsed_data.get("projects", [])
    if projects_items:
        projects_score = min(10, len(projects_items) * 3)
        feedback.append(f"✓ Projects section found with {len(projects_items)} projects")
    else:
        feedback.append("⚠ No projects section found")

    scoring_breakdown["Projects"] = {"score": projects_score, "max": 10}
    score += projects_score

    cert_score = 0
    cert_items = parsed_data.get("certifications", [])
    if cert_items:
        cert_score = min(8, len(cert_items) * 2)
        feedback.append(f"✓ Certifications found: {len(cert_items)} certificates")
    else:
        feedback.append("⚠ No certifications listed")

    scoring_breakdown["Certifications"] = {"score": cert_score, "max": 8}
    score += cert_score

    research_score = 0
    research_items = parsed_data.get("research", [])
    if research_items:
        research_score = min(7, len(research_items) * 3)
        feedback.append(f"✓ Research/Publications found: {len(research_items)} papers")

    scoring_breakdown["Research & Publications"] = {"score": research_score, "max": 7}
    score += research_score

    formatting_score = 0
    section_count = len([k for k, v in parsed_data.items() if v])
    if section_count >= 4:
        formatting_score += 5
        feedback.append("✓ Well-structured with multiple sections")

    action_verbs = [
        "developed",
        "designed",
        "implemented",
        "managed",
        "led",
        "created",
        "built",
        "improved",
        "optimized",
        "achieved",
        "delivered",
        "launched",
    ]
    text_lower = full_text.lower()
    found_verbs = sum(1 for verb in action_verbs if verb in text_lower)
    if found_verbs >= 5:
        formatting_score += 5
        feedback.append(f"✓ Strong action verbs used ({found_verbs} found)")
    elif found_verbs >= 2:
        formatting_score += 3
        feedback.append(f"⚠ Some action verbs used ({found_verbs} found)")

    scoring_breakdown["Formatting & Keywords"] = {"score": formatting_score, "max": 10}
    score += formatting_score

    percentage = round((score / max_score) * 100, 1)

    if percentage >= 90:
        rating = "Excellent"
        rating_color = "green"
    elif percentage >= 75:
        rating = "Very Good"
        rating_color = "blue"
    elif percentage >= 60:
        rating = "Good"
        rating_color = "yellow"
    elif percentage >= 40:
        rating = "Fair"
        rating_color = "orange"
    else:
        rating = "Needs Improvement"
        rating_color = "red"

    recommendations = []
    if contact_score < 10:
        recommendations.append("Add complete contact information (email and phone)")
    if education_score < 10:
        recommendations.append("Add or expand your education section")
    if experience_score < 15:
        recommendations.append("Add more work experience details with quantifiable achievements")
    if skills_score < 10:
        recommendations.append("Expand your skills section with relevant technical and soft skills")
    if projects_score < 5:
        recommendations.append("Include relevant projects to showcase your abilities")
    if cert_score < 4:
        recommendations.append("Add professional certifications if available")
    if formatting_score < 7:
        recommendations.append("Use more action verbs and improve resume structure")

    return {
        "score": score,
        "max_score": max_score,
        "percentage": percentage,
        "rating": rating,
        "rating_color": rating_color,
        "feedback": feedback,
        "scoring_breakdown": scoring_breakdown,
        "recommendations": recommendations
        if recommendations
        else ["Your resume looks great! Keep it updated."],
    }

# =====================================================
# 🔐 AUTHENTICATION ROUTES
# =====================================================

@app.route("/api/auth/signup", methods=["POST"])
def signup():
    try:
        data = request.get_json()
        print("📩 Signup data received:", data)

        required = ["email", "password", "name", "userType"]
        missing = [f for f in required if not data.get(f)]
        if missing:
            return jsonify({"message": f"Missing fields: {', '.join(missing)}"}), 400

        email = data["email"].lower().strip()
        password = data["password"]
        name = data["name"].strip()
        userType = data["userType"].strip()

        valid_user_types = ["student", "admin", "placementCell"]
        if userType not in valid_user_types:
            return jsonify({"message": f"Invalid userType. Must be one of: {', '.join(valid_user_types)}"}), 400

        if users_collection.find_one({"email": email}):
            return jsonify({"message": "User already exists"}), 400

        hashed_pw = generate_password_hash(password)
        new_user = {
            "_id": str(uuid.uuid4()),
            "email": email,
            "password": hashed_pw,
            "name": name,
            "userType": userType,
            "onboardingCompleted": False if userType == "student" else True,
            "createdAt": datetime.utcnow()
        }
        users_collection.insert_one(new_user)

        token = generate_token(new_user)
        print("✅ User created successfully:", new_user["_id"], "Type:", userType)

        user_data = {
            "id": str(new_user["_id"]),
            "email": new_user["email"],
            "name": new_user.get("name", ""),
            "userType": new_user.get("userType", "student")
        }

        return jsonify({
            "message": "Signup successful!",
            "token": token,
            "user": user_data
        }), 201

    except Exception as e:
        print("❌ Signup error:", e)
        return jsonify({"message": "Internal server error"}), 500

@app.route("/api/auth/login", methods=["POST"])
def login():
    try:
        data = request.get_json()
        print("📩 Login data received:", data)

        if not data.get("email") or not data.get("password"):
            return jsonify({"message": "Email and password are required"}), 400

        email = data["email"].lower().strip()
        password = data["password"]

        user = users_collection.find_one({"email": email})
        if not user or not check_password_hash(user["password"], password):
            return jsonify({"message": "Invalid email or password"}), 401

        token = generate_token(user)
        print("✅ Login successful for:", email, "Type:", user.get("userType"))

        user_data = {
            "id": str(user["_id"]),
            "email": user["email"],
            "name": user["name"],
            "userType": user["userType"]
        }

        return jsonify({
            "message": "Login successful!",
            "token": token,
            "user": user_data
        }), 200

    except Exception as e:
        print("❌ Login error:", e)
        return jsonify({"message": f"Internal server error: {str(e)}"}), 500

@app.route("/api/auth/status", methods=["GET"])
def status():
    """Validate token and return authenticated status"""
    current_user = get_current_user()
    if not current_user:
        return jsonify({"authenticated": False}), 401

    return jsonify({
        "authenticated": True,
        "user": {
            "id": str(current_user["_id"]),
            "email": current_user["email"],
            "name": current_user["name"],
            "userType": current_user["userType"]
        }
    }), 200

# =====================================================
# 👤 USER PROFILE ROUTES
# =====================================================

@app.route("/api/user", methods=["GET"])
def get_user_profile():
    """Get current user's profile"""
    current_user = get_current_user()
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    user_data = {
        "id": str(current_user["_id"]),
        "email": current_user["email"],
        "name": current_user.get("name", ""),
        "userType": current_user.get("userType", "student"),
        "field": current_user.get("field", ""),
        "year": current_user.get("year", ""),
        "cgpa": current_user.get("cgpa", 0),
        "mobile": current_user.get("mobile", ""),
        "resumeUrl": current_user.get("resumeUrl", ""),
        "performanceDocUrl": current_user.get("performanceDocUrl", ""),
        "onboardingCompleted": current_user.get("onboardingCompleted", False),
        "rollNo": current_user.get("rollNo", ""),
        "skills": current_user.get("skills", []),
        "techStack": current_user.get("techStack", []),
        "aiTools": current_user.get("aiTools", []),
        "experiences": current_user.get("experiences", []),
        "certifications": current_user.get("certifications", []),
        "projects": current_user.get("projects", [])
    }

    return jsonify(user_data), 200

@app.route("/api/user", methods=["PUT"])
def update_user_profile():
    """Update current user's profile"""
    current_user = get_current_user()
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    try:
        data = request.get_json()
        
        # Fields that can be updated
        update_fields = {}
        if "name" in data:
            update_fields["name"] = data["name"]
        if "field" in data:
            update_fields["field"] = data["field"]
        if "year" in data:
            update_fields["year"] = data["year"]
        if "cgpa" in data:
            update_fields["cgpa"] = float(data["cgpa"])
        if "mobile" in data:
            update_fields["mobile"] = data["mobile"]
        if "resumeUrl" in data:
            update_fields["resumeUrl"] = data["resumeUrl"]
        if "performanceDocUrl" in data:
            update_fields["performanceDocUrl"] = data["performanceDocUrl"]
        if "rollNo" in data:
            update_fields["rollNo"] = data["rollNo"]
        if "skills" in data:
            update_fields["skills"] = data["skills"]
        if "techStack" in data:
            update_fields["techStack"] = data["techStack"]
        if "aiTools" in data:
            update_fields["aiTools"] = data["aiTools"]
        if "experiences" in data:
            update_fields["experiences"] = data["experiences"]
        if "certifications" in data:
            update_fields["certifications"] = data["certifications"]
        if "projects" in data:
            update_fields["projects"] = data["projects"]
        if "onboardingCompleted" in data:
            update_fields["onboardingCompleted"] = data["onboardingCompleted"]

        users_collection.update_one(
            {"_id": str(current_user["_id"])},
            {"$set": update_fields}
        )

        return jsonify({"message": "Profile updated successfully"}), 200

    except Exception as e:
        print(f"❌ Error updating profile: {e}")
        return jsonify({"message": "Failed to update profile"}), 500

# =====================================================
# 🚀 ONBOARDING ROUTES
# =====================================================

@app.route("/api/onboarding", methods=["POST", "OPTIONS"])
def onboarding():
    """Handle multi-step student onboarding and save data to user profile"""
    # Handle OPTIONS preflight request (CORS)
    if request.method == "OPTIONS":
        return jsonify({"message": "OK"}), 200
    
    current_user = get_current_user()
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    try:
        data = request.get_json() or {}
        step = data.get("step")
        payload = data.get("data", {})
        completed = data.get("completed", False)
        

        update_fields = {}

        # Step 1: Basic Info (branch, year, mobile, cgpa, rollNo)
        if step == "basic_info":
            update_fields.update({
                "field": payload.get("field"),
                "year": payload.get("year"),
                "mobile": payload.get("mobile"),
                "cgpa": payload.get("cgpa"),
                "rollNo": payload.get("rollNo"),
            })
            print(f"✅ Onboarding Step 1 - Basic Info: {update_fields}")

        # Step 2: Experiences (work experience + LinkedIn)
        elif step == "experiences":
            update_fields["experiences"] = payload.get("experiences", [])
            if payload.get("linkedinProfile"):
                update_fields["linkedinProfile"] = payload.get("linkedinProfile")
            print(f"✅ Onboarding Step 2 - Experiences: {len(update_fields.get('experiences', []))} items")

        # Step 3: Certifications/Achievements
        elif step == "certifications":
            update_fields["certifications"] = payload.get("achievements", [])
            print(f"✅ Onboarding Step 3 - Certifications: {len(update_fields.get('certifications', []))} items")

        # Step 4: Projects (projects + GitHub profile)
        elif step == "projects":
            update_fields["projects"] = payload.get("projects", [])
            if payload.get("githubProfile"):
                update_fields["githubProfile"] = payload.get("githubProfile")
            print(f"✅ Onboarding Step 4 - Projects: {len(update_fields.get('projects', []))} items")

        # Step 5: Skills (area of expertise)
        elif step == "skills":
            update_fields["skills"] = payload.get("skills", [])
            print(f"✅ Onboarding Step 5 - Skills: {update_fields.get('skills', [])}")

        # Step 6: Tech Stack (programming languages & tools)
        elif step == "tech_stack":
            update_fields["techStack"] = payload.get("techStack", [])
            print(f"✅ Onboarding Step 6 - Tech Stack: {update_fields.get('techStack', [])}")

        # Step 7: AI Tools (final step)
        elif step == "ai_tools":
            update_fields["aiTools"] = payload.get("aiTools", [])
            print(f"✅ Onboarding Step 7 - AI Tools: {update_fields.get('aiTools', [])}")

        # Mark onboarding as complete if specified
        if completed:
            update_fields["onboardingCompleted"] = True
            print(f"✅ Onboarding marked as COMPLETED for user: {current_user['email']}")

        # Update user document in database
        if update_fields:
            users_collection.update_one(
                {"_id": str(current_user["_id"])},
                {"$set": update_fields}
            )
            print(f"✅ Database updated for user: {current_user['email']}")

        return jsonify({"message": "Onboarding data saved successfully"}), 200

    except Exception as e:
        print(f"❌ Error in onboarding route: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"message": "Failed to save onboarding data"}), 500


@app.route("/api/onboarding/status", methods=["GET"])
def onboarding_status():
    """Return onboarding completion status for the current user"""
    current_user = get_current_user()
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    return jsonify({
        "onboardingCompleted": current_user.get("onboardingCompleted", False)
    }), 200


# =====================================================
# � ATS RESUME UPLOAD & SCORING
# =====================================================

@app.route("/api/ats/upload", methods=["POST"])
def upload_resume_for_ats():
    current_user = get_current_user()
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files["file"]

    if file.filename == "":
        return jsonify({"error": "No file selected"}), 400

    filename = file.filename.lower()

    try:
        if filename.endswith(".pdf"):
            text = extract_text_from_pdf(file)
        elif filename.endswith(".docx"):
            text = extract_text_from_docx(file)
        elif filename.endswith(".txt"):
            text = file.read().decode("utf-8", errors="ignore")
        else:
            return jsonify({"error": "Unsupported file format. Please upload PDF, DOCX, or TXT"}), 400

        parsed_data = parse_resume(text)
        ats_score = calculate_ats_score(parsed_data, text)

        return jsonify(
            {
                "success": True,
                "filename": file.filename,
                "parsed_data": parsed_data,
                "ats_score": ats_score,
            }
        ), 200

    except Exception as e:
        print(f"Error processing ATS resume: {str(e)}")
        return jsonify({"error": f"Error processing file: {str(e)}"}), 500

@app.route("/api/ats/from_saved_resume", methods=["POST"])
def ats_from_saved_resume():
    """Run ATS analysis using the current user's stored resumeUrl (Google Drive link)."""
    current_user = get_current_user()
    if not current_user:
        return jsonify({"error": "Not authenticated"}), 401

    resume_url = current_user.get("resumeUrl", "").strip()
    if not resume_url:
        return jsonify({"error": "No resume URL found for user"}), 400

    if not validate_drive_link(resume_url):
        return jsonify({"error": "Invalid resume URL. Must be a Google Drive or Docs link"}), 400

    try:
        # Convert common Google Drive share URLs to a direct download URL
        direct_url = resume_url
        if "drive.google.com/file/d/" in resume_url:
            match = re.search(r"/d/([^/]+)", resume_url)
            if match:
                file_id = match.group(1)
                direct_url = f"https://drive.google.com/uc?export=download&id={file_id}"

        response = requests.get(direct_url, stream=True, timeout=30)
        if response.status_code != 200:
            return jsonify({"error": "Failed to download resume from Drive"}), 502

        content_type = response.headers.get("Content-Type", "").lower()
        text = ""

        # Try to infer format from content-type or URL
        if ".pdf" in resume_url or "application/pdf" in content_type:
            pdf_bytes = BytesIO(response.content)
            text = extract_text_from_pdf(pdf_bytes)
        elif ".docx" in resume_url or "application/vnd.openxmlformats-officedocument.wordprocessingml.document" in content_type:
            docx_bytes = BytesIO(response.content)
            text = extract_text_from_docx(docx_bytes)
        else:
            # Fallback: treat as plain text
            try:
                text = response.content.decode("utf-8", errors="ignore")
            except Exception:
                return jsonify({"error": "Unsupported resume format. Please use PDF, DOCX or text"}), 400

        if not text.strip():
            return jsonify({"error": "Could not extract text from resume"}), 400

        parsed_data = parse_resume(text)
        ats_score = calculate_ats_score(parsed_data, text)

        return jsonify({
            "parsed_data": parsed_data,
            "ats_score": ats_score
        }), 200

    except Exception as e:
        print(f"❌ Error in ATS from saved resume: {e}")
        return jsonify({"error": "Failed to analyze resume from saved link"}), 500

# =====================================================
# �� STUDENT'S PERSONAL PROJECTS (Their own projects)
# =====================================================

@app.route("/api/user/projects", methods=["GET"])
def get_student_personal_projects():
    """Get student's own projects (not admin opportunities)"""
    current_user = get_current_user()
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    try:
        projects = list(project_collection.find({
            "userId": str(current_user["_id"])
        }).sort("createdAt", -1))

        for project in projects:
            project["id"] = str(project["_id"])
            del project["_id"]

        return jsonify({"projects": projects}), 200

    except Exception as e:
        print(f"❌ Error fetching personal projects: {e}")
        return jsonify({"message": "Failed to fetch projects"}), 500

@app.route("/api/user/projects", methods=["POST"])
def create_student_personal_project():
    """Student creates their own project"""
    current_user = get_current_user()
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    try:
        data = request.get_json()

        if not data.get("title"):
            return jsonify({"message": "Title is required"}), 400

        project = {
            "_id": str(uuid.uuid4()),
            "userId": str(current_user["_id"]),
            "title": data["title"],
            "githubLink": data.get("githubLink", ""),
            "websiteLink": data.get("websiteLink", ""),
            "techStack": data.get("techStack", []),
            "createdAt": datetime.utcnow(),
            "updatedAt": datetime.utcnow()
        }

        project_collection.insert_one(project)
        project["id"] = project["_id"]
        del project["_id"]

        return jsonify({
            "message": "Project created successfully",
            "project": project
        }), 201

    except Exception as e:
        print(f"❌ Error creating project: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"message": f"Failed to create project: {str(e)}"}), 500

@app.route("/api/user/projects/<project_id>", methods=["PUT"])
def update_student_personal_project(project_id):
    """Student updates their own project"""
    current_user = get_current_user()
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    try:
        data = request.get_json()

        # Verify ownership
        project = project_collection.find_one({
            "_id": project_id,
            "userId": str(current_user["_id"])
        })

        if not project:
            return jsonify({"message": "Project not found"}), 404

        update_fields = {
            "title": data.get("title"),
            "githubLink": data.get("githubLink"),
            "websiteLink": data.get("websiteLink"),
            "techStack": data.get("techStack"),
            "updatedAt": datetime.utcnow()
        }

        update_fields = {k: v for k, v in update_fields.items() if v is not None}

        project_collection.update_one(
            {"_id": project_id},
            {"$set": update_fields}
        )

        return jsonify({"message": "Project updated successfully"}), 200

    except Exception as e:
        print(f"❌ Error updating project: {e}")
        return jsonify({"message": "Failed to update project"}), 500

@app.route("/api/user/projects/<project_id>", methods=["DELETE"])
def delete_student_personal_project(project_id):
    """Student deletes their own project"""
    current_user = get_current_user()
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    try:
        result = project_collection.delete_one({
            "_id": project_id,
            "userId": str(current_user["_id"])
        })

        if result.deleted_count == 0:
            return jsonify({"message": "Project not found"}), 404

        return jsonify({"message": "Project deleted successfully"}), 200

    except Exception as e:
        print(f"❌ Error deleting project: {e}")
        return jsonify({"message": "Failed to delete project"}), 500

# =====================================================
# 🎓 ADMIN OPPORTUNITY ROUTES - Projects
# =====================================================

@app.route("/api/user/admin/projects", methods=["GET"])
def get_admin_projects():
    """Get all admin-created project opportunities"""
    current_user = get_current_user()
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    try:
        projects = list(admin_projects_collection.find().sort("createdAt", -1))

        for project in projects:
            project["id"] = str(project["_id"])

            # Get application count
            app_count = student_applications_collection.count_documents({
                "opportunityId": project["id"],
                "opportunityType": "project"
            })
            project["applicationCount"] = app_count

            del project["_id"]

        return jsonify({"projects": projects}), 200

    except Exception as e:
        print(f"❌ Error fetching admin projects: {e}")
        return jsonify({"message": "Failed to fetch projects"}), 500

@app.route("/api/user/admin/projects", methods=["POST"])
def create_admin_project():
    """Create new admin project opportunity"""
    current_user = get_current_user()
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    if current_user.get("userType") != "admin":
        return jsonify({"message": "Unauthorized"}), 403

    try:
        data = request.get_json()

        required_fields = ["title", "domain", "studentsRequired", "googleFormLink"]
        for field in required_fields:
            if field not in data:
                return jsonify({"message": f"Missing required field: {field}"}), 400

        project = {
            "_id": str(uuid.uuid4()),
            "title": data["title"],
            "domain": data["domain"],
            "studentsRequired": int(data["studentsRequired"]),
            "duration": data.get("duration", ""),
            "deadline": data.get("deadline", ""),
            "googleFormLink": data["googleFormLink"],
            "description": data.get("description", ""),
            "requirements": data.get("requirements", ""),
            "professors": data.get("professors", []),
            "students": data.get("students", []),
            "createdBy": str(current_user["_id"]),
            "createdAt": datetime.utcnow(),
            "updatedAt": datetime.utcnow(),
            "status": "active"
        }

        admin_projects_collection.insert_one(project)
        project["id"] = project["_id"]
        del project["_id"]
        project["applicationCount"] = 0

        return jsonify({"message": "Project created successfully", "project": project}), 201

    except Exception as e:
        print(f"❌ Error creating admin project: {e}")
        return jsonify({"message": "Failed to create project"}), 500

@app.route("/api/user/admin/projects/<project_id>", methods=["PUT"])
def update_admin_project(project_id):
    """Update admin project opportunity"""
    current_user = get_current_user()
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    if current_user.get("userType") != "admin":
        return jsonify({"message": "Unauthorized"}), 403

    try:
        data = request.get_json()

        update_fields = {
            "title": data.get("title"),
            "domain": data.get("domain"),
            "studentsRequired": data.get("studentsRequired"),
            "duration": data.get("duration"),
            "deadline": data.get("deadline"),
            "googleFormLink": data.get("googleFormLink"),
            "description": data.get("description"),
            "requirements": data.get("requirements"),
            "professors": data.get("professors"),
            "students": data.get("students"),
            "status": data.get("status"),
            "updatedAt": datetime.utcnow()
        }

        update_fields = {k: v for k, v in update_fields.items() if v is not None}

        result = admin_projects_collection.update_one(
            {"_id": project_id},
            {"$set": update_fields}
        )

        if result.matched_count == 0:
            return jsonify({"message": "Project not found"}), 404

        return jsonify({"message": "Project updated successfully"}), 200

    except Exception as e:
        print(f"❌ Error updating admin project: {e}")
        return jsonify({"message": "Failed to update project"}), 500

@app.route("/api/user/admin/projects/<project_id>", methods=["DELETE"])
def delete_admin_project(project_id):
    """Delete admin project opportunity"""
    current_user = get_current_user()
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    if current_user.get("userType") != "admin":
        return jsonify({"message": "Unauthorized"}), 403

    try:
        result = admin_projects_collection.delete_one({"_id": project_id})

        if result.deleted_count == 0:
            return jsonify({"message": "Project not found"}), 404

        return jsonify({"message": "Project deleted successfully"}), 200

    except Exception as e:
        print(f"❌ Error deleting admin project: {e}")
        return jsonify({"message": "Failed to delete project"}), 500

# =====================================================
# 📚 ADMIN OPPORTUNITY ROUTES - Research
# =====================================================

@app.route("/api/user/admin/research", methods=["GET"])
def get_admin_research():
    """Get all admin-created research opportunities"""
    current_user = get_current_user()
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    try:
        papers = list(admin_research_collection.find().sort("createdAt", -1))

        for paper in papers:
            paper["id"] = str(paper["_id"])

            app_count = student_applications_collection.count_documents({
                "opportunityId": paper["id"],
                "opportunityType": "research"
            })
            paper["applicationCount"] = app_count

            del paper["_id"]

        return jsonify({"papers": papers}), 200

    except Exception as e:
        print(f"❌ Error fetching admin research: {e}")
        return jsonify({"message": "Failed to fetch research papers"}), 500

@app.route("/api/user/admin/research", methods=["POST"])
def create_admin_research():
    """Create new admin research opportunity"""
    current_user = get_current_user()
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    if current_user.get("userType") != "admin":
        return jsonify({"message": "Unauthorized"}), 403

    try:
        data = request.get_json()

        required_fields = ["title", "domain", "studentsRequired", "googleFormLink"]
        for field in required_fields:
            if field not in data:
                return jsonify({"message": f"Missing required field: {field}"}), 400

        paper = {
            "_id": str(uuid.uuid4()),
            "title": data["title"],
            "domain": data["domain"],
            "studentsRequired": int(data["studentsRequired"]),
            "duration": data.get("duration", ""),
            "deadline": data.get("deadline", ""),
            "googleFormLink": data["googleFormLink"],
            "description": data.get("description", ""),
            "requirements": data.get("requirements", ""),
            "professors": data.get("professors", []),
            "students": data.get("students", []),
            "createdBy": str(current_user["_id"]),
            "createdAt": datetime.utcnow(),
            "updatedAt": datetime.utcnow(),
            "status": "active"
        }

        admin_research_collection.insert_one(paper)
        paper["id"] = paper["_id"]
        del paper["_id"]
        paper["applicationCount"] = 0

        return jsonify({"message": "Research paper created successfully", "paper": paper}), 201

    except Exception as e:
        print(f"❌ Error creating admin research: {e}")
        return jsonify({"message": "Failed to create research paper"}), 500

@app.route("/api/user/admin/research/<paper_id>", methods=["PUT"])
def update_admin_research(paper_id):
    """Update admin research opportunity"""
    current_user = get_current_user()
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    if current_user.get("userType") != "admin":
        return jsonify({"message": "Unauthorized"}), 403

    try:
        data = request.get_json()

        update_fields = {
            "title": data.get("title"),
            "domain": data.get("domain"),
            "studentsRequired": data.get("studentsRequired"),
            "duration": data.get("duration"),
            "deadline": data.get("deadline"),
            "googleFormLink": data.get("googleFormLink"),
            "description": data.get("description"),
            "requirements": data.get("requirements"),
            "professors": data.get("professors"),
            "students": data.get("students"),
            "status": data.get("status"),
            "updatedAt": datetime.utcnow()
        }

        update_fields = {k: v for k, v in update_fields.items() if v is not None}

        result = admin_research_collection.update_one(
            {"_id": paper_id},
            {"$set": update_fields}
        )

        if result.matched_count == 0:
            return jsonify({"message": "Research paper not found"}), 404

        return jsonify({"message": "Research paper updated successfully"}), 200

    except Exception as e:
        print(f"❌ Error updating admin research: {e}")
        return jsonify({"message": "Failed to update research paper"}), 500

@app.route("/api/user/admin/research/<paper_id>", methods=["DELETE"])
def delete_admin_research(paper_id):
    """Delete admin research opportunity"""
    current_user = get_current_user()
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    if current_user.get("userType") != "admin":
        return jsonify({"message": "Unauthorized"}), 403

    try:
        result = admin_research_collection.delete_one({"_id": paper_id})

        if result.deleted_count == 0:
            return jsonify({"message": "Research paper not found"}), 404

        return jsonify({"message": "Research paper deleted successfully"}), 200

    except Exception as e:
        print(f"❌ Error deleting admin research: {e}")
        return jsonify({"message": "Failed to delete research paper"}), 500

# =====================================================
# 🔬 ADMIN OPPORTUNITY ROUTES - Patents
# =====================================================

@app.route("/api/user/admin/patents", methods=["GET"])
def get_admin_patents():
    """Get all admin-created patent opportunities"""
    current_user = get_current_user()
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    try:
        patents = list(admin_patents_collection.find().sort("createdAt", -1))

        for patent in patents:
            patent["id"] = str(patent["_id"])

            app_count = student_applications_collection.count_documents({
                "opportunityId": patent["id"],
                "opportunityType": "patent"
            })
            patent["applicationCount"] = app_count

            del patent["_id"]

        return jsonify({"patents": patents}), 200

    except Exception as e:
        print(f"❌ Error fetching admin patents: {e}")
        return jsonify({"message": "Failed to fetch patents"}), 500

@app.route("/api/user/admin/patents", methods=["POST"])
def create_admin_patent():
    """Create new admin patent opportunity"""
    current_user = get_current_user()
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    if current_user.get("userType") != "admin":
        return jsonify({"message": "Unauthorized"}), 403

    try:
        data = request.get_json()

        required_fields = ["title", "domain", "studentsRequired", "googleFormLink"]
        for field in required_fields:
            if field not in data:
                return jsonify({"message": f"Missing required field: {field}"}), 400

        patent = {
            "_id": str(uuid.uuid4()),
            "title": data["title"],
            "domain": data["domain"],
            "studentsRequired": int(data["studentsRequired"]),
            "duration": data.get("duration", ""),
            "deadline": data.get("deadline", ""),
            "googleFormLink": data["googleFormLink"],
            "description": data.get("description", ""),
            "requirements": data.get("requirements", ""),
            "professors": data.get("professors", []),
            "students": data.get("students", []),
            "createdBy": str(current_user["_id"]),
            "createdAt": datetime.utcnow(),
            "updatedAt": datetime.utcnow(),
            "status": "active"
        }

        admin_patents_collection.insert_one(patent)
        patent["id"] = patent["_id"]
        del patent["_id"]
        patent["applicationCount"] = 0

        return jsonify({"message": "Patent created successfully", "patent": patent}), 201

    except Exception as e:
        print(f"❌ Error creating admin patent: {e}")
        return jsonify({"message": "Failed to create patent"}), 500

@app.route("/api/user/admin/patents/<patent_id>", methods=["PUT"])
def update_admin_patent(patent_id):
    """Update admin patent opportunity"""
    current_user = get_current_user()
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    if current_user.get("userType") != "admin":
        return jsonify({"message": "Unauthorized"}), 403

    try:
        data = request.get_json()

        update_fields = {
            "title": data.get("title"),
            "domain": data.get("domain"),
            "studentsRequired": data.get("studentsRequired"),
            "duration": data.get("duration"),
            "deadline": data.get("deadline"),
            "googleFormLink": data.get("googleFormLink"),
            "description": data.get("description"),
            "requirements": data.get("requirements"),
            "professors": data.get("professors"),
            "students": data.get("students"),
            "status": data.get("status"),
            "updatedAt": datetime.utcnow()
        }

        update_fields = {k: v for k, v in update_fields.items() if v is not None}

        result = admin_patents_collection.update_one(
            {"_id": patent_id},
            {"$set": update_fields}
        )

        if result.matched_count == 0:
            return jsonify({"message": "Patent not found"}), 404

        return jsonify({"message": "Patent updated successfully"}), 200

    except Exception as e:
        print(f"❌ Error updating admin patent: {e}")
        return jsonify({"message": "Failed to update patent"}), 500

@app.route("/api/user/admin/patents/<patent_id>", methods=["DELETE"])
def delete_admin_patent(patent_id):
    """Delete admin patent opportunity"""
    current_user = get_current_user()
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    if current_user.get("userType") != "admin":
        return jsonify({"message": "Unauthorized"}), 403

    try:
        result = admin_patents_collection.delete_one({"_id": patent_id})

        if result.deleted_count == 0:
            return jsonify({"message": "Patent not found"}), 404

        return jsonify({"message": "Patent deleted successfully"}), 200

    except Exception as e:
        print(f"❌ Error deleting admin patent: {e}")
        return jsonify({"message": "Failed to delete patent"}), 500

# =====================================================
# 👥 ADMIN ROUTES - Student Management
# =====================================================

@app.route("/api/admin/students", methods=["GET"])
def get_all_students():
    """Get all students for admin view"""
    current_user = get_current_user()
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    if current_user.get("userType") != "admin":
        return jsonify({"message": "Unauthorized"}), 403

    try:
        # Fetch all users with userType = "student"
        students = list(users_collection.find(
            {"userType": "student"}
        ).sort("createdAt", -1))

        formatted_students = []
        for student in students:
            # Get student's personal projects count
            projects_count = project_collection.count_documents({"userId": str(student["_id"])})
            
            # Get student's applications count
            applications_count = student_applications_collection.count_documents({"studentId": str(student["_id"])})

            formatted_student = {
                "id": str(student["_id"]),
                "name": student.get("name", ""),
                "email": student.get("email", ""),
                "field": student.get("field", ""),
                "year": student.get("year", ""),
                "mobile": student.get("mobile", ""),
                "cgpa": student.get("cgpa", 0),
                "rollNo": student.get("rollNo", ""),
                "resumeUrl": student.get("resumeUrl", ""),
                "skills": student.get("skills", []),
                "techStack": student.get("techStack", []),
                "aiTools": student.get("aiTools", []),
                "experiences": student.get("experiences", []),
                "certifications": student.get("certifications", []),
                "projects": student.get("projects", []),  # Onboarding projects
                "projectsCount": projects_count,  # Personal projects count
                "applicationsCount": applications_count,
                "onboardingCompleted": student.get("onboardingCompleted", False),
                "createdAt": student.get("createdAt").isoformat() if isinstance(student.get("createdAt"), datetime) else student.get("createdAt")
            }
            formatted_students.append(formatted_student)

        return jsonify({"students": formatted_students}), 200

    except Exception as e:
        print(f"❌ Error fetching students: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"message": "Failed to fetch students"}), 500

# =====================================================
# 🆘 HELP / SUPPORT ROUTES
# =====================================================

@app.route("/api/help/reports", methods=["POST"])
def create_help_report():
    """Create a new help/support report from any authenticated user."""
    current_user = get_current_user()
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    try:
        data = request.get_json() or {}
        title = (data.get("title") or "").strip()
        description = (data.get("description") or "").strip()

        if not title or not description:
            return jsonify({"message": "Title and description are required"}), 400

        user_type = current_user.get("userType", "unknown")

        report = {
            "_id": str(uuid.uuid4()),
            "userId": str(current_user.get("_id")),
            "userEmail": current_user.get("email", ""),
            "userName": current_user.get("name", ""),
            "userType": user_type,
            "title": title,
            "description": description,
            "createdAt": datetime.utcnow(),
            "status": "open",
        }

        help_reports_collection.insert_one(report)

        return jsonify({
            "message": "Help report submitted successfully",
            "id": report["_id"],
        }), 201

    except Exception as e:
        print(f"❌ Error creating help report: {e}")
        return jsonify({"message": "Failed to submit help report"}), 500

# =====================================================
# 📝 STUDENT ROUTES - View Opportunities
# =====================================================

@app.route("/api/student/opportunities/projects", methods=["GET"])
def get_student_project_opportunities():
    """Get all active project opportunities for students to apply to"""
    current_user = get_current_user()
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    try:
        projects = list(admin_projects_collection.find(
            {"status": "active"}
        ).sort("createdAt", -1))

        for project in projects:
            project["id"] = str(project["_id"])

            # Check if current student has already applied
            existing_application = student_applications_collection.find_one({
                "studentId": str(current_user["_id"]),
                "opportunityId": project["id"],
                "opportunityType": "project"
            })
            project["hasApplied"] = existing_application is not None
            project["applicationStatus"] = existing_application.get("status") if existing_application else None

            del project["_id"]

        return jsonify({"projects": projects}), 200

    except Exception as e:
        print(f"❌ Error fetching student project opportunities: {e}")
        return jsonify({"message": "Failed to fetch project opportunities"}), 500

@app.route("/api/student/opportunities/research", methods=["GET"])
def get_student_research_opportunities():
    """Get all active research opportunities for students to apply to"""
    current_user = get_current_user()
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    try:
        papers = list(admin_research_collection.find(
            {"status": "active"}
        ).sort("createdAt", -1))

        for paper in papers:
            paper["id"] = str(paper["_id"])

            existing_application = student_applications_collection.find_one({
                "studentId": str(current_user["_id"]),
                "opportunityId": paper["id"],
                "opportunityType": "research"
            })
            paper["hasApplied"] = existing_application is not None
            paper["applicationStatus"] = existing_application.get("status") if existing_application else None

            del paper["_id"]

        return jsonify({"papers": papers}), 200

    except Exception as e:
        print(f"❌ Error fetching student research opportunities: {e}")
        return jsonify({"message": "Failed to fetch research opportunities"}), 500

@app.route("/api/student/opportunities/patents", methods=["GET"])
def get_student_patent_opportunities():
    """Get all active patent opportunities for students to apply to"""
    current_user = get_current_user()
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    try:
        patents = list(admin_patents_collection.find(
            {"status": "active"}
        ).sort("createdAt", -1))

        for patent in patents:
            patent["id"] = str(patent["_id"])

            existing_application = student_applications_collection.find_one({
                "studentId": str(current_user["_id"]),
                "opportunityId": patent["id"],
                "opportunityType": "patent"
            })
            patent["hasApplied"] = existing_application is not None
            patent["applicationStatus"] = existing_application.get("status") if existing_application else None

            del patent["_id"]

        return jsonify({"patents": patents}), 200

    except Exception as e:
        print(f"❌ Error fetching student patent opportunities: {e}")
        return jsonify({"message": "Failed to fetch patent opportunities"}), 500

# =====================================================
# 📝 STUDENT ROUTES - Submit Applications
# =====================================================

@app.route("/api/student/applications", methods=["POST"])
def submit_student_application():
    """Student submits application to an opportunity with drive links"""
    current_user = get_current_user()
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    if current_user.get("userType") != "student":
        return jsonify({"message": "Only students can submit applications"}), 403

    try:
        data = request.get_json()

        # Validate required fields
        required_fields = ["opportunityId", "opportunityType", "resumeLink", "submissionLink"]
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({"message": f"Missing required field: {field}"}), 400

        opportunity_id = data["opportunityId"]
        opportunity_type = data["opportunityType"]

        # Validate opportunity type
        if opportunity_type not in ["project", "research", "patent"]:
            return jsonify({"message": "Invalid opportunity type"}), 400

        # Check if opportunity exists
        if opportunity_type == "project":
            opportunity = admin_projects_collection.find_one({"_id": opportunity_id})
        elif opportunity_type == "research":
            opportunity = admin_research_collection.find_one({"_id": opportunity_id})
        else:  # patent
            opportunity = admin_patents_collection.find_one({"_id": opportunity_id})

        if not opportunity:
            return jsonify({"message": "Opportunity not found"}), 404

        # Check if student already applied
        existing_application = student_applications_collection.find_one({
            "studentId": str(current_user["_id"]),
            "opportunityId": opportunity_id,
            "opportunityType": opportunity_type
        })

        if existing_application:
            return jsonify({"message": "You have already applied to this opportunity"}), 400

        # Validate Google Drive links
        resume_link = data["resumeLink"].strip()
        submission_link = data["submissionLink"].strip()

        if not validate_drive_link(resume_link):
            return jsonify({"message": "Invalid resume link. Must be a Google Drive link"}), 400

        if not validate_drive_link(submission_link):
            return jsonify({"message": "Invalid submission link. Must be a Google Drive link"}), 400

        # Create application
        application = {
            "_id": str(uuid.uuid4()),
            "studentId": str(current_user["_id"]),
            "studentName": current_user.get("name", "Unknown"),
            "studentEmail": current_user.get("email", ""),
            "studentBranch": current_user.get("field", "Not specified"),
            "studentYear": current_user.get("year", "Not specified"),
            "studentCGPA": current_user.get("cgpa", 0),
            "opportunityId": opportunity_id,
            "opportunityType": opportunity_type,
            "opportunityTitle": opportunity.get("title", ""),
            "resumeLink": resume_link,
            "submissionLink": submission_link,
            "additionalLinks": data.get("additionalLinks", []),
            "coverLetter": data.get("coverLetter", ""),
            "status": "pending",
            "appliedAt": datetime.utcnow(),
            "updatedAt": datetime.utcnow(),
            "adminNotes": ""
        }

        student_applications_collection.insert_one(application)
        application["id"] = str(application["_id"])
        del application["_id"]

        return jsonify({
            "message": "Application submitted successfully",
            "application": application
        }), 201

    except Exception as e:
        print(f"❌ Error submitting application: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"message": f"Failed to submit application: {str(e)}"}), 500

@app.route("/api/student/applications", methods=["GET"])
def get_student_applications():
    """Get all applications submitted by current student"""
    current_user = get_current_user()
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    if current_user.get("userType") != "student":
        return jsonify({"message": "Only students can view their applications"}), 403

    try:
        applications = list(student_applications_collection.find({
            "studentId": str(current_user["_id"])
        }).sort("appliedAt", -1))

        for app in applications:
            app["id"] = str(app["_id"])
            del app["_id"]

        return jsonify({"applications": applications}), 200

    except Exception as e:
        print(f"❌ Error fetching applications: {e}")
        return jsonify({"message": "Failed to fetch applications"}), 500

@app.route("/api/student/applications/<application_id>", methods=["PUT"])
def update_student_application(application_id):
    """Student updates their pending application"""
    current_user = get_current_user()
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    if current_user.get("userType") != "student":
        return jsonify({"message": "Only students can update their applications"}), 403

    try:
        # Find application
        application = student_applications_collection.find_one({
            "_id": application_id,
            "studentId": str(current_user["_id"])
        })

        if not application:
            return jsonify({"message": "Application not found"}), 404

        # Can only update pending applications
        if application.get("status") != "pending":
            return jsonify({"message": "Can only update pending applications"}), 400

        data = request.get_json()

        # Fields that can be updated
        update_fields = {}
        if "resumeLink" in data:
            update_fields["resumeLink"] = data["resumeLink"].strip()
        if "submissionLink" in data:
            update_fields["submissionLink"] = data["submissionLink"].strip()
        if "additionalLinks" in data:
            update_fields["additionalLinks"] = data["additionalLinks"]
        if "coverLetter" in data:
            update_fields["coverLetter"] = data["coverLetter"]

        update_fields["updatedAt"] = datetime.utcnow()

        student_applications_collection.update_one(
            {"_id": application_id},
            {"$set": update_fields}
        )

        return jsonify({"message": "Application updated successfully"}), 200

    except Exception as e:
        print(f"❌ Error updating application: {e}")
        return jsonify({"message": "Failed to update application"}), 500

@app.route("/api/student/applications/<application_id>", methods=["DELETE"])
def delete_student_application(application_id):
    """Student withdraws their pending application"""
    current_user = get_current_user()
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    if current_user.get("userType") != "student":
        return jsonify({"message": "Only students can delete their applications"}), 403

    try:
        # Find application
        application = student_applications_collection.find_one({
            "_id": application_id,
            "studentId": str(current_user["_id"])
        })

        if not application:
            return jsonify({"message": "Application not found"}), 404

        # Can only delete pending applications
        if application.get("status") != "pending":
            return jsonify({"message": "Can only withdraw pending applications"}), 400

        student_applications_collection.delete_one({"_id": application_id})

        return jsonify({"message": "Application withdrawn successfully"}), 200

    except Exception as e:
        print(f"❌ Error deleting application: {e}")
        return jsonify({"message": "Failed to withdraw application"}), 500

# =====================================================
# 👨‍💼 ADMIN ROUTES - View & Manage Applications
# =====================================================

@app.route("/api/admin/applications", methods=["GET"])
def get_all_applications():
    """Admin gets all applications with optional filters"""
    current_user = get_current_user()
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    if current_user.get("userType") != "admin":
        return jsonify({"message": "Unauthorized. Admin access required."}), 403

    try:
        # Get query parameters for filtering
        opportunity_type = request.args.get("opportunityType")
        status = request.args.get("status")
        opportunity_id = request.args.get("opportunityId")

        # Build query
        query = {}
        if opportunity_type:
            query["opportunityType"] = opportunity_type
        if status:
            query["status"] = status
        if opportunity_id:
            query["opportunityId"] = opportunity_id

        applications = list(student_applications_collection.find(query).sort("appliedAt", -1))

        for app in applications:
            app["id"] = str(app["_id"])
            del app["_id"]

        return jsonify({
            "applications": applications,
            "total": len(applications)
        }), 200

    except Exception as e:
        print(f"❌ Error fetching applications: {e}")
        return jsonify({"message": "Failed to fetch applications"}), 500

@app.route("/api/admin/applications/<application_id>/status", methods=["PUT"])
def update_application_status(application_id):
    """Admin approves or rejects an application"""
    current_user = get_current_user()
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    if current_user.get("userType") != "admin":
        return jsonify({"message": "Unauthorized. Admin access required."}), 403

    try:
        data = request.get_json()
        new_status = data.get("status")
        admin_notes = data.get("adminNotes", "")

        if new_status not in ["approved", "rejected", "pending"]:
            return jsonify({"message": "Invalid status. Must be 'approved', 'rejected', or 'pending'"}), 400

        result = student_applications_collection.update_one(
            {"_id": application_id},
            {
                "$set": {
                    "status": new_status,
                    "adminNotes": admin_notes,
                    "reviewedBy": str(current_user["_id"]),
                    "reviewedAt": datetime.utcnow(),
                    "updatedAt": datetime.utcnow()
                }
            }
        )

        if result.matched_count == 0:
            return jsonify({"message": "Application not found"}), 404

        return jsonify({"message": f"Application {new_status} successfully"}), 200

    except Exception as e:
        print(f"❌ Error updating application status: {e}")
        return jsonify({"message": "Failed to update application status"}), 500

@app.route("/api/admin/opportunities/<opportunity_id>/applications", methods=["GET"])
def get_opportunity_applications(opportunity_id):
    """Get all applications for a specific opportunity"""
    current_user = get_current_user()
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    if current_user.get("userType") != "admin":
        return jsonify({"message": "Unauthorized. Admin access required."}), 403

    try:
        opportunity_type = request.args.get("type", "project")

        applications = list(student_applications_collection.find({
            "opportunityId": opportunity_id,
            "opportunityType": opportunity_type
        }).sort("appliedAt", -1))

        for app in applications:
            app["id"] = str(app["_id"])
            del app["_id"]

        # Get stats
        total = len(applications)
        pending = len([a for a in applications if a.get("status") == "pending"])
        approved = len([a for a in applications if a.get("status") == "approved"])
        rejected = len([a for a in applications if a.get("status") == "rejected"])

        return jsonify({
            "applications": applications,
            "stats": {
                "total": total,
                "pending": pending,
                "approved": approved,
                "rejected": rejected
            }
        }), 200

    except Exception as e:
        print(f"❌ Error fetching opportunity applications: {e}")
        return jsonify({"message": "Failed to fetch applications"}), 500

@app.route("/api/admin/applications/export", methods=["GET"])
def export_applications():
    """Export applications to Excel file"""
    current_user = get_current_user()
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    if current_user.get("userType") != "admin":
        return jsonify({"message": "Unauthorized. Admin access required."}), 403

    try:
        # Get query parameters
        opportunity_type = request.args.get("opportunityType")
        status = request.args.get("status")

        query = {}
        if opportunity_type:
            query["opportunityType"] = opportunity_type
        if status:
            query["status"] = status

        applications = list(student_applications_collection.find(query).sort("appliedAt", -1))

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Applications"

        # Headers
        headers = [
            "Application ID", "Student Name", "Email", "Branch", "Year", "CGPA",
            "Opportunity Type", "Opportunity Title", "Resume Link", "Submission Link",
            "Status", "Applied Date", "Admin Notes"
        ]

        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data rows
        for row_num, app in enumerate(applications, 2):
            ws.cell(row=row_num, column=1, value=app.get("_id", ""))
            ws.cell(row=row_num, column=2, value=app.get("studentName", ""))
            ws.cell(row=row_num, column=3, value=app.get("studentEmail", ""))
            ws.cell(row=row_num, column=4, value=app.get("studentBranch", ""))
            ws.cell(row=row_num, column=5, value=str(app.get("studentYear", "")))
            ws.cell(row=row_num, column=6, value=app.get("studentCGPA", 0))
            ws.cell(row=row_num, column=7, value=app.get("opportunityType", ""))
            ws.cell(row=row_num, column=8, value=app.get("opportunityTitle", ""))
            ws.cell(row=row_num, column=9, value=app.get("resumeLink", ""))
            ws.cell(row=row_num, column=10, value=app.get("submissionLink", ""))
            ws.cell(row=row_num, column=11, value=app.get("status", ""))
            ws.cell(row=row_num, column=12, value=app.get("appliedAt", "").strftime("%Y-%m-%d %H:%M") if app.get("appliedAt") else "")
            ws.cell(row=row_num, column=13, value=app.get("adminNotes", ""))

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'applications_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )

    except Exception as e:
        print(f"❌ Error exporting applications: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"message": "Failed to export applications"}), 500

# =====================================================
# 🤖 ADMIN AI CHAT ROUTES
# =====================================================



@app.route("/api/ai/admin/chat", methods=["POST", "OPTIONS"])
def admin_ai_chat():
    """Admin AI chat to query and filter students using natural language"""
    # Handle OPTIONS preflight request (CORS)
    if request.method == "OPTIONS":
        return jsonify({"message": "OK"}), 200
    
    current_user = get_current_user()
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    if current_user.get("userType") != "admin":
        return jsonify({"message": "Unauthorized. Admin access required."}), 403

    try:
        data = request.get_json()
        user_message = data.get("message", "").strip()

        if not user_message:
            return jsonify({"error": "Message is required"}), 400

        # Get all students from database
        all_students = list(users_collection.find({"userType": "student"}))
        
        # Convert ObjectId to string and prepare student data
        students_data = []
        for student in all_students:
            student_info = {
                "id": str(student["_id"]),
                "name": student.get("name", ""),
                "email": student.get("email", ""),
                "field": student.get("field", ""),
                "year": student.get("year", ""),
                "cgpa": student.get("cgpa", 0),
                "rollNo": student.get("rollNo", ""),
                "skills": student.get("skills", []),
                "techStack": student.get("techStack", []),
                "aiTools": student.get("aiTools", []),
                "experiences": student.get("experiences", []),
                "certifications": student.get("certifications", []),
                "projects": student.get("projects", []),
                "linkedinProfile": student.get("linkedinProfile", ""),
                "githubProfile": student.get("githubProfile", ""),
                "mobile": student.get("mobile", "")
            }
            students_data.append(student_info)

        # Use Gemini AI to process the query and filter students
        client = genai.Client()

        # Create a prompt for Gemini to understand the query and filter students
        prompt = f"""You are an AI assistant helping an admin filter students based on their query.

User Query: "{user_message}"

Available Students Data (JSON):
{students_data}

Instructions:
1. Analyze the user's query to understand what type of students they're looking for
2. Filter the students based on skills, techStack, aiTools, experiences, field, year, CGPA, or any other criteria mentioned
3. Return a JSON response with:
   - "response": A natural language response explaining what you found (2-3 sentences max)
   - "filtered_student_ids": An array of student IDs that match the criteria

Example queries:
- "Show me students with Blockchain skills" → Filter by skills containing "Blockchain"
- "Show me MERN stack developers" → Filter by techStack containing MongoDB, Express, React, Node
- "Show me AI/ML students" → Filter by skills/techStack containing AI, ML, Machine Learning, etc.
- "Show fullstack developers with CGPA above 8" → Filter by skills AND cgpa

Return ONLY valid JSON in this exact format:
{{
  "response": "Found X students with [criteria]. They have experience in [relevant skills].",
  "filtered_student_ids": ["id1", "id2", "id3"]
}}

Be intelligent about matching - use synonyms and related terms. For example:
- "AI" matches "Artificial Intelligence", "Machine Learning", "Deep Learning"
- "MERN" matches "MongoDB", "Express", "React", "Node.js"
- "Fullstack" matches students with both frontend and backend skills
"""

        print(f"🔍 Calling Gemini API with model: gemini-2.5-flash")
        print(f"🔍 Number of students: {len(students_data)}")
        
        response = client.models.generate_content(
            model='gemini-2.5-flash',
            contents=prompt
        )

        # Parse Gemini response
        ai_text = response.text.strip()
        print(f"🤖 Gemini raw response (first 500 chars): {ai_text[:500]}")
        
        # Extract JSON from response (handle markdown code blocks)
        if "```json" in ai_text:
            ai_text = ai_text.split("```json")[1].split("```")[0].strip()
        elif "```" in ai_text:
            ai_text = ai_text.split("```")[1].split("```")[0].strip()

        print(f"📝 Extracted JSON: {ai_text[:300]}")
        
        import json
        ai_response = json.loads(ai_text)
        print(f"✅ JSON parsed successfully")

        # Get filtered students based on IDs returned by AI
        filtered_student_ids = ai_response.get("filtered_student_ids", [])
        filtered_students = [s for s in students_data if s["id"] in filtered_student_ids]

        print(f"✅ Found {len(filtered_students)} students")
        
        return jsonify({
            "response": ai_response.get("response", "Here are the students matching your criteria."),
            "students": filtered_students
        }), 200

    except Exception as e:
        error_msg = str(e)
        print(f"❌ Error in admin AI chat: {error_msg}")
        print(f"❌ Error type: {type(e).__name__}")
        import traceback
        traceback.print_exc()
        
        return jsonify({
            "response": f"Error: {error_msg}",
            "students": [],
            "error_type": type(e).__name__
        }), 200


@app.route("/api/ai/admin/export-filtered-students", methods=["POST", "OPTIONS"])
def export_filtered_students():
    """Export filtered students to Excel"""
    # Handle OPTIONS preflight request (CORS)
    if request.method == "OPTIONS":
        return jsonify({"message": "OK"}), 200
    
    current_user = get_current_user()
    if not current_user:
        return jsonify({"message": "Not authenticated"}), 401

    if current_user.get("userType") != "admin":
        return jsonify({"message": "Unauthorized. Admin access required."}), 403

    try:
        data = request.get_json()
        students = data.get("students", [])

        if not students:
            return jsonify({"error": "No students to export"}), 400

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Filtered Students"

        # Headers
        headers = [
            "Name", "Email", "Roll No", "Branch", "Year", "CGPA",
            "Mobile", "Skills", "Tech Stack", "AI Tools",
            "LinkedIn", "GitHub", "Experience Count", "Projects Count", "Certifications Count"
        ]

        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data rows
        for row_num, student in enumerate(students, 2):
            ws.cell(row=row_num, column=1, value=student.get("name", ""))
            ws.cell(row=row_num, column=2, value=student.get("email", ""))
            ws.cell(row=row_num, column=3, value=student.get("rollNo", ""))
            ws.cell(row=row_num, column=4, value=student.get("field", ""))
            ws.cell(row=row_num, column=5, value=student.get("year", ""))
            ws.cell(row=row_num, column=6, value=student.get("cgpa", 0))
            ws.cell(row=row_num, column=7, value=student.get("mobile", ""))
            ws.cell(row=row_num, column=8, value=", ".join(student.get("skills", [])))
            ws.cell(row=row_num, column=9, value=", ".join(student.get("techStack", [])))
            ws.cell(row=row_num, column=10, value=", ".join(student.get("aiTools", [])))
            ws.cell(row=row_num, column=11, value=student.get("linkedinProfile", ""))
            ws.cell(row=row_num, column=12, value=student.get("githubProfile", ""))
            ws.cell(row=row_num, column=13, value=len(student.get("experiences", [])))
            ws.cell(row=row_num, column=14, value=len(student.get("projects", [])))
            ws.cell(row=row_num, column=15, value=len(student.get("certifications", [])))

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"filtered_students_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

    except Exception as e:
        print(f"❌ Error exporting filtered students: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": "Failed to export students"}), 500


# =====================================================
# 🏃 RUN SERVER
# =====================================================
if __name__ == "__main__":
    print("🚀 Starting Flask server with Socket.IO...")
    print("=" * 60)
    print("✅ Authentication system enabled")
    print("✅ User profile management enabled")
    print("✅ Student personal projects enabled")
    print("✅ Admin opportunity management enabled")
    print("✅ Student application system enabled")
    print("✅ Admin application review enabled")
    print("=" * 60)
    print("🌐 Server running on http://localhost:5001")
    print("=" * 60)

    socketio.run(app, host="0.0.0.0", port=5001, debug=True)